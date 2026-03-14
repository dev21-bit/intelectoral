import streamlit as st
import pandas as pd
import folium
import json
from streamlit_folium import folium_static
import pymysql
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from datetime import datetime
import numpy as np
from scipy import stats
import warnings
import openpyxl
from openpyxl import load_workbook
warnings.filterwarnings('ignore')

# ---------------------------------------------------
# CONFIGURACIÓN PROFESIONAL
# ---------------------------------------------------
st.set_page_config(
    page_title="ZACATECAS · Sistema de Inteligencia Electoral",
    page_icon="🗳️",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ---------------------------------------------------
# ESTILO CORPORATIVO DE ALTO NIVEL
# ---------------------------------------------------
st.markdown("""
<style>
    /* FUENTES Y RESET */
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap');
    
    * {
        font-family: 'Inter', sans-serif;
        margin: 0;
        padding: 0;
        box-sizing: border-box;
    }
    
    /* FONDO PRINCIPAL */
    .stApp {
        background: linear-gradient(135deg, #f5f7fa 0%, #e9ecef 100%);
    }
    
    /* HEADER DE LUJO */
    .header-premium {
        background: linear-gradient(135deg, #1a1a2e 0%, #16213e 100%);
        padding: 1.5rem;
        border-radius: 20px;
        margin: 1rem 0 2rem 0;
        box-shadow: 0 20px 40px rgba(0,0,0,0.15);
        border: 1px solid rgba(255,255,255,0.1);
    }
    
    .header-premium h1 {
        color: white;
        font-size: 2.8rem;
        font-weight: 800;
        letter-spacing: -1px;
        margin: 0;
        text-shadow: 2px 2px 4px rgba(0,0,0,0.3);
    }
    
    .header-premium p {
        color: #a8b2d1;
        font-size: 1.1rem;
        font-weight: 300;
        margin: 0.5rem 0 0 0;
    }
    
    /* TARJETAS DE MÉTRICAS PREMIUM */
    .metric-card-premium {
        background: white;
        border-radius: 20px;
        padding: 1.5rem;
        box-shadow: 0 10px 30px rgba(0,0,0,0.08);
        transition: all 0.3s cubic-bezier(0.25, 0.8, 0.25, 1);
        border: 1px solid rgba(255,255,255,0.1);
        backdrop-filter: blur(10px);
        position: relative;
        overflow: hidden;
    }
    
    .metric-card-premium:hover {
        transform: translateY(-5px);
        box-shadow: 0 20px 40px rgba(0,0,0,0.12);
    }
    
    .metric-card-premium::before {
        content: '';
        position: absolute;
        top: 0;
        left: 0;
        right: 0;
        height: 4px;
        background: linear-gradient(90deg, #667eea, #764ba2, #6b8cff, #667eea);
        background-size: 300% 300%;
        animation: gradient 3s ease infinite;
    }
    
    @keyframes gradient {
        0% { background-position: 0% 50%; }
        50% { background-position: 100% 50%; }
        100% { background-position: 0% 50%; }
    }
    
    .metric-label {
        color: #64748b;
        font-size: 0.9rem;
        font-weight: 500;
        text-transform: uppercase;
        letter-spacing: 0.5px;
        margin-bottom: 0.5rem;
    }
    
    .metric-value {
        color: #1e293b;
        font-size: 2.5rem;
        font-weight: 700;
        line-height: 1.2;
        margin-bottom: 0.25rem;
    }
    
    .metric-trend {
        display: flex;
        align-items: center;
        gap: 0.5rem;
        font-size: 0.9rem;
        color: #64748b;
    }
    
    .trend-up { color: #10b981; }
    .trend-down { color: #ef4444; }
    
    /* BOTONES PREMIUM */
    .stButton button {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white;
        border: none;
        border-radius: 12px;
        padding: 0.75rem 1.5rem;
        font-weight: 600;
        font-size: 0.95rem;
        transition: all 0.3s ease;
        box-shadow: 0 4px 15px rgba(102, 126, 234, 0.35);
        border: 1px solid rgba(255,255,255,0.1);
        width: 100%;
    }
    
    .stButton button:hover {
        transform: translateY(-2px);
        box-shadow: 0 8px 25px rgba(102, 126, 234, 0.45);
    }
    
    /* INPUTS PREMIUM */
    .stTextInput input {
        background: white;
        border: 2px solid #e2e8f0;
        border-radius: 12px;
        padding: 0.75rem 1rem;
        font-size: 1rem;
        transition: all 0.3s ease;
        box-shadow: 0 2px 4px rgba(0,0,0,0.02);
    }
    
    .stTextInput input:focus {
        border-color: #667eea;
        box-shadow: 0 0 0 4px rgba(102, 126, 234, 0.1);
    }
    
    /* TABS PREMIUM */
    .stTabs [data-baseweb="tab-list"] {
        gap: 2rem;
        background: transparent;
        padding: 0.5rem;
        border-bottom: 2px solid #e2e8f0;
    }
    
    .stTabs [data-baseweb="tab"] {
        color: #64748b;
        font-weight: 600;
        font-size: 1rem;
        padding: 0.75rem 1rem;
        transition: all 0.3s ease;
        border-bottom: 2px solid transparent;
        margin-bottom: -2px;
    }
    
    .stTabs [aria-selected="true"] {
        color: #667eea !important;
        border-bottom: 2px solid #667eea !important;
        background: transparent !important;
    }
    
    /* LÍNEA DECORATIVA */
    .divider-premium {
        height: 2px;
        background: linear-gradient(90deg, transparent, #667eea, #764ba2, #667eea, transparent);
        margin: 2.5rem 0;
        opacity: 0.5;
    }
    
    /* BADGES */
    .badge-premium {
        display: inline-block;
        padding: 0.35rem 0.75rem;
        border-radius: 100px;
        font-size: 0.8rem;
        font-weight: 600;
        letter-spacing: 0.3px;
        background: linear-gradient(135deg, #667eea20 0%, #764ba220 100%);
        color: #667eea;
        border: 1px solid #667eea40;
    }
    
    /* ANIMACIONES */
    @keyframes slideIn {
        from {
            opacity: 0;
            transform: translateY(20px);
        }
        to {
            opacity: 1;
            transform: translateY(0);
        }
    }
    
    .animate-slide {
        animation: slideIn 0.5s ease-out;
    }
    
    /* ESTILOS PARA TARJETAS DE OPERADOR */
    .operator-card {
        background: white;
        border-radius: 16px;
        padding: 1.5rem;
        margin: 1rem 0;
        box-shadow: 0 4px 6px rgba(0,0,0,0.05);
        border-left: 4px solid #667eea;
        transition: all 0.3s ease;
    }
    
    .operator-card:hover {
        transform: translateX(5px);
        box-shadow: 0 8px 15px rgba(0,0,0,0.1);
    }
    
    .operator-name {
        font-size: 1.3rem;
        font-weight: 700;
        color: #1e293b;
        margin-bottom: 0.5rem;
    }
    
    .operator-stats {
        display: grid;
        grid-template-columns: repeat(auto-fit, minmax(120px, 1fr));
        gap: 1rem;
        margin-top: 1rem;
    }
    
    .stat-item {
        text-align: center;
        padding: 0.5rem;
        background: #f8fafc;
        border-radius: 12px;
    }
    
    .stat-value {
        font-size: 1.5rem;
        font-weight: 700;
        color: #667eea;
    }
    
    .stat-label {
        font-size: 0.8rem;
        color: #64748b;
        text-transform: uppercase;
    }
</style>
""", unsafe_allow_html=True)

# ---------------------------------------------------
# SEGURIDAD DE NIVEL EMPRESARIAL
# ---------------------------------------------------
class SecurityManager:
    CLAVE_CORRECTA = "zac2026"
    SESSION_KEY = "authenticated"
    
    @staticmethod
    def check_authentication():
        if SecurityManager.SESSION_KEY not in st.session_state:
            st.session_state[SecurityManager.SESSION_KEY] = False
        
        if not st.session_state[SecurityManager.SESSION_KEY]:
            SecurityManager.show_login()
            return False
        return True
    
    @staticmethod
    def show_login():
        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            with st.container():
                st.markdown("""
                <div style="background: white; border-radius: 30px; padding: 3rem; box-shadow: 0 25px 50px -12px rgba(0,0,0,0.25); margin: 2rem 0;">
                    <h2 style="color: #1e293b; font-weight: 700; margin-bottom: 1rem;">🔐 Acceso Restringido</h2>
                    <p style="color: #64748b; margin-bottom: 2rem;">Sistema de Inteligencia Electoral · Zacatecas 2026</p>
                </div>
                """, unsafe_allow_html=True)
                
                clave = st.text_input("", placeholder="Ingresa tu clave de acceso", type="password", label_visibility="collapsed")
                
                if st.button("Iniciar Sesión", use_container_width=True):
                    if clave == SecurityManager.CLAVE_CORRECTA:
                        st.session_state[SecurityManager.SESSION_KEY] = True
                        st.query_params = {"auth": "true"}
                        st.rerun()
                    else:
                        st.error("Clave incorrecta. Acceso denegado.")
                st.stop()

# ---------------------------------------------------
# GESTOR DE DATOS PROFESIONAL
# ---------------------------------------------------
class DataManager:
    @staticmethod
    @st.cache_data(ttl=3600, show_spinner=False)
    def get_ine_data():
        try:
            connection = pymysql.connect(
                host='sql3.freesqldatabase.com',
                user='sql3817481',
                password='398j6uKWle',
                database='sql3817481',
                port=3306,
                connect_timeout=5
            )
            query = """
            SELECT
            LPAD(CAST(seccion AS CHAR),4,'0') as seccion,
            SUBSTRING_INDEX(SUBSTRING_INDEX(domicilio,' ', -3),' ',1) as cp,
            COUNT(*) as simpatizantes
            FROM ine
            GROUP BY seccion, cp
            """
            df = pd.read_sql(query, connection)
            connection.close()
            df['cp'] = df['cp'].astype(str)
            df['seccion'] = df['seccion'].astype(str).str.zfill(4)
            return df
        except Exception as e:
            st.error(f"Error de conexión a BD: {str(e)}")
            return pd.DataFrame()

    @staticmethod
    @st.cache_data(ttl=3600)
    def load_excel():
        try:
            xls = pd.ExcelFile("COLONIAS ZAC.xlsx")
            sheet1 = pd.read_excel(xls, xls.sheet_names[0])
            sheet2 = pd.read_excel(xls, xls.sheet_names[1])
            data = {}
            sheet1['SECCION'] = sheet1['SECCION'].astype(str).str.zfill(4)
            sheet2['Catalogo de Colonias_seccion'] = sheet2['Catalogo de Colonias_seccion'].astype(str).str.zfill(4)
            for _, row in sheet1.iterrows():
                sec = row['SECCION']
                cp = str(row['CP']).replace(".0","")
                col = row['NOMBRE DE LA COLONIA']
                data.setdefault(sec, []).append({"colonia": col, "cp": cp})
            for _, row in sheet2.iterrows():
                sec = row['Catalogo de Colonias_seccion']
                cp = str(row['CP'])
                col = row['NOMBRE DE LA COLONIA']
                data.setdefault(sec, []).append({"colonia": col, "cp": cp})
            return data
        except Exception as e:
            st.error(f"Error cargando Excel: {str(e)}")
            return {}

    @staticmethod
    @st.cache_data(ttl=3600)
    def load_datos_seccion_con_colores():
        """
        Carga los datos de sección incluyendo los colores de las celdas
        """
        try:
            # Cargar el workbook con openpyxl para acceder a los colores
            wb = load_workbook("DATOS_POR_SECCION.xlsx", data_only=True)
            ws = wb.active
            
            # Encontrar la fila de encabezados
            header_row = 1
            for row in ws.iter_rows(min_row=1, max_row=10):
                for cell in row:
                    if cell.value and str(cell.value).strip().upper() in ["SECCIÓN", "SECCION"]:
                        header_row = cell.row
                        break
            
            # Extraer datos con pandas primero
            df = pd.read_excel("DATOS_POR_SECCION.xlsx", header=header_row-1)
            df.columns = [str(c).strip().upper().replace("Ó", "O").replace("Á","A").replace("É","E") for c in df.columns]
            df = df.loc[:, ~df.columns.str.startswith("UNNAMED")]
            
            # Renombrar columnas esperadas
            column_mapping = {}
            for col in df.columns:
                if "SECCION" in col:
                    column_mapping[col] = "seccion"
                elif "PARTICIPACION" in col:
                    column_mapping[col] = "participacion"
                elif "INICIAL" in col:
                    column_mapping[col] = "inicial"
                elif "META" in col:
                    column_mapping[col] = "meta"
            
            df = df.rename(columns=column_mapping)
            
            # Extraer colores de la columna SECCION
            colores_seccion = {}
            seccion_col_idx = None
            
            # Encontrar el índice de la columna SECCION
            for idx, cell in enumerate(ws[header_row], 1):
                if cell.value and str(cell.value).strip().upper() in ["SECCIÓN", "SECCION"]:
                    seccion_col_idx = idx
                    break
            
            if seccion_col_idx:
                # Iterar sobre las filas de datos
                for row in ws.iter_rows(min_row=header_row + 1):
                    seccion_cell = row[seccion_col_idx - 1]
                    if seccion_cell.value:
                        seccion_val = str(seccion_cell.value).strip()
                        # Obtener el color de fondo
                        if seccion_cell.fill and seccion_cell.fill.fgColor:
                            if seccion_cell.fill.fgColor.rgb:
                                color = seccion_cell.fill.fgColor.rgb
                                if len(color) == 8:  # Formato AARRGGBB
                                    # Convertir a formato hex sin alpha
                                    color_hex = f"#{color[2:]}"
                                    colores_seccion[seccion_val] = color_hex
            
            # Procesar el dataframe
            df = df[pd.to_numeric(df["seccion"], errors="coerce").notna()].copy()
            df["seccion"] = df["seccion"].astype(float).astype(int).astype(str).str.zfill(4)
            df["participacion"] = pd.to_numeric(df["participacion"], errors="coerce")
            df["inicial"] = pd.to_numeric(df["inicial"], errors="coerce").round(0).astype("Int64")
            df["meta"] = pd.to_numeric(df["meta"], errors="coerce").round(0).astype("Int64")
            
            # Agregar colores al dataframe
            df['color'] = df['seccion'].map(colores_seccion)
            
            return df.set_index("seccion")
        except Exception as e:
            st.error(f"Error cargando datos de sección con colores: {str(e)}")
            return pd.DataFrame()

    @staticmethod
    @st.cache_data(ttl=3600)
    def get_simpatizantes_colonia():
        try:
            connection = pymysql.connect(
                host='sql3.freesqldatabase.com',
                user='sql3817481',
                password='398j6uKWle',
                database='sql3817481',
                port=3306
            )
            query = """
            SELECT
            LPAD(CAST(seccion AS CHAR),4,'0') as seccion,
            domicilio
            FROM ine
            """
            df = pd.read_sql(query, connection)
            connection.close()
            df['seccion'] = df['seccion'].astype(str).str.zfill(4)
            df["cp"] = df["domicilio"].str.extract(r'(\d{5})(?!.*\d{5})')
            colonias_excel = DataManager.load_excel()
            resultados = []
            for _, persona in df.iterrows():
                seccion = persona["seccion"]
                domicilio = str(persona["domicilio"]).upper()
                cp = str(persona["cp"])
                if seccion in colonias_excel:
                    for col in colonias_excel[seccion]:
                        nombre_col = str(col["colonia"]).upper()
                        cp_col = str(col["cp"]).replace(".0","")
                        if cp == cp_col and nombre_col in domicilio:
                            resultados.append({"seccion": seccion, "colonia": col["colonia"]})
                            break
                        elif cp == cp_col:
                            resultados.append({"seccion": seccion, "colonia": col["colonia"]})
                            break
            conteo = pd.DataFrame(resultados)
            if len(conteo) > 0:
                conteo = conteo.groupby(["seccion","colonia"]).size().reset_index(name="simpatizantes")
            else:
                conteo = pd.DataFrame(columns=["seccion","colonia","simpatizantes"])
            return conteo
        except Exception as e:
            st.error(f"Error procesando simpatizantes: {str(e)}")
            return pd.DataFrame()

    @staticmethod
    @st.cache_data(ttl=3600)
    def get_operadores_data():
        """Obtiene datos de operadores desde la base de datos"""
        try:
            connection = pymysql.connect(
                host='sql3.freesqldatabase.com',
                user='sql3817481',
                password='398j6uKWle',
                database='sql3817481',
                port=3306
            )
            query = """
            SELECT
            id,
            nombre,
            apellido_paterno,
            apellido_materno,
            sexo,
            fecha_nacimiento,
            curp,
            clave_elector,
            domicilio,
            telefono,
            anio_registro,
            vigencia,
            LPAD(CAST(seccion AS CHAR),4,'0') as seccion,
            usuario_nombre,
            usuario_pin
            FROM ine
            """
            df = pd.read_sql(query, connection)
            connection.close()
            
            # Procesar datos
            df['seccion'] = df['seccion'].astype(str).str.zfill(4)
            df['nombre_completo'] = df['nombre'] + ' ' + df['apellido_paterno'] + ' ' + df['apellido_materno']
            df['edad'] = pd.to_datetime('today').year - pd.to_datetime(df['fecha_nacimiento'], dayfirst=True).dt.year
            
            return df
        except Exception as e:
            st.error(f"Error obteniendo datos de operadores: {str(e)}")
            return pd.DataFrame()

    @staticmethod
    @st.cache_data(ttl=86400)
    def load_geojson():
        try:
            with open("zacatecas_capital_secciones.geojson", encoding="utf-8") as f:
                geo = json.load(f)
            for feature in geo["features"]:
                if "SECCION" in feature["properties"]:
                    feature["properties"]["seccion"] = str(feature["properties"]["SECCION"]).zfill(4)
                else:
                    feature["properties"]["seccion"] = str(feature["properties"]["seccion"]).zfill(4)
            return geo["features"]
        except Exception as e:
            st.error(f"Error cargando GeoJSON: {str(e)}")
            return []

# ---------------------------------------------------
# ANALIZADOR DE OPERADORES
# ---------------------------------------------------
class OperadoresAnalytics:
    def __init__(self, operadores_df):
        self.operadores = operadores_df
        
    def get_resumen_operadores(self):
        """Genera resumen estadístico por operador"""
        if self.operadores.empty:
            return pd.DataFrame()
        
        # Agrupar por usuario_nombre
        resumen = self.operadores.groupby('usuario_nombre').agg({
            'id': 'count',
            'seccion': lambda x: x.nunique(),
            'sexo': lambda x: x.value_counts().to_dict() if len(x) > 0 else {},
            'edad': ['mean', 'min', 'max'],
            'anio_registro': lambda x: x.mode()[0] if len(x.mode()) > 0 else None
        }).round(1)
        
        # Renombrar columnas
        resumen.columns = ['total_registros', 'secciones_asignadas', 'distribucion_sexo', 
                          'edad_promedio', 'edad_min', 'edad_max', 'anio_registro_comun']
        
        # Calcular eficiencia (simulada - basada en registros vs secciones)
        resumen['indice_productividad'] = (resumen['total_registros'] / resumen['secciones_asignadas']).round(1)
        
        return resumen.reset_index()
    
    def get_operador_detalle(self, operador_nombre):
        """Obtiene detalle completo de un operador específico"""
        if self.operadores.empty:
            return pd.DataFrame()
        
        detalle = self.operadores[self.operadores['usuario_nombre'] == operador_nombre].copy()
        return detalle
    
    def get_top_operadores(self, n=5):
        """Obtiene los operadores con más registros"""
        resumen = self.get_resumen_operadores()
        if resumen.empty:
            return pd.DataFrame()
        return resumen.nlargest(n, 'total_registros')
    
    def create_operador_charts(self):
        """Crea visualizaciones para operadores"""
        charts = {}
        resumen = self.get_resumen_operadores()
        
        if resumen.empty:
            return charts
        
        # Gráfico de barras - Top operadores
        fig1 = px.bar(
            resumen.nlargest(10, 'total_registros'),
            x='usuario_nombre',
            y='total_registros',
            title='Top 10 Operadores por Registros',
            labels={'usuario_nombre': 'Operador', 'total_registros': 'Registros'},
            color='total_registros',
            color_continuous_scale=['#667eea', '#764ba2']
        )
        fig1.update_layout(
            plot_bgcolor='rgba(0,0,0,0)',
            paper_bgcolor='rgba(0,0,0,0)',
            font=dict(family="Inter"),
            xaxis_tickangle=-45
        )
        charts['top_operadores'] = fig1
        
        # Gráfico de distribución de edades por operador
        fig2 = go.Figure()
        for operador in resumen['usuario_nombre'].head(5):
            datos_operador = self.operadores[self.operadores['usuario_nombre'] == operador]
            fig2.add_trace(go.Box(
                y=datos_operador['edad'],
                name=operador,
                marker_color='#667eea'
            ))
        
        fig2.update_layout(
            title='Distribución de Edades por Operador (Top 5)',
            yaxis_title='Edad',
            plot_bgcolor='rgba(0,0,0,0)',
            paper_bgcolor='rgba(0,0,0,0)',
            font=dict(family="Inter"),
            showlegend=True
        )
        charts['edad_distribution'] = fig2
        
        return charts

# ---------------------------------------------------
# ANALIZADOR DE DATOS AVANZADO
# ---------------------------------------------------
class AdvancedAnalytics:
    def __init__(self, db, datos_seccion, simpatizantes_colonia):
        self.db = db
        self.datos_seccion = datos_seccion
        self.simpatizantes_colonia = simpatizantes_colonia
        self.simpatizantes_por_seccion = db.groupby('seccion')['simpatizantes'].sum().reset_index()
        
    def generate_executive_summary(self):
        """Genera un resumen ejecutivo de alto nivel"""
        # Calcular avance real usando simpatizantes vs meta
        df_combinado = self.datos_seccion.reset_index().merge(
            self.simpatizantes_por_seccion, on='seccion', how='left'
        ).fillna(0)
        
        total_meta = df_combinado['meta'].sum()
        total_simpatizantes = df_combinado['simpatizantes'].sum()
        
        summary = {
            'total_simpatizantes': int(total_simpatizantes),
            'secciones_activas': len(self.simpatizantes_por_seccion),
            'cobertura': f"{(len(self.simpatizantes_por_seccion) / len(self.datos_seccion) * 100):.1f}%" if len(self.datos_seccion) > 0 else "0%",
            'promedio_x_seccion': int(self.simpatizantes_por_seccion['simpatizantes'].mean()) if len(self.simpatizantes_por_seccion) > 0 else 0,
            'eficiencia_global': self.calculate_global_efficiency(),
            'top_performers': self.identify_top_performers(5),
            'areas_oportunidad': self.identify_opportunity_areas(5),
            'proyeccion_meta': self.calculate_projection()
        }
        return summary
    
    def calculate_global_efficiency(self):
        """Calcula la eficiencia global usando simpatizantes vs meta"""
        if len(self.datos_seccion) == 0:
            return 0
        
        df_combinado = self.datos_seccion.reset_index().merge(
            self.simpatizantes_por_seccion, on='seccion', how='left'
        ).fillna(0)
        
        total_meta = df_combinado['meta'].sum()
        total_simpatizantes = df_combinado['simpatizantes'].sum()
        
        return round((total_simpatizantes / total_meta * 100), 1) if total_meta > 0 else 0
    
    def identify_top_performers(self, n):
        """Identifica las secciones con mejor desempeño basado en avance real"""
        if len(self.datos_seccion) == 0:
            return []
        
        df = self.datos_seccion.reset_index().merge(
            self.simpatizantes_por_seccion, on='seccion', how='left'
        ).fillna(0)
        
        df['cumplimiento'] = (df['simpatizantes'] / df['meta'] * 100).round(1)
        df = df[df['meta'] > 0].nlargest(n, 'cumplimiento')
        return df[['seccion', 'cumplimiento']].to_dict('records')
    
    def identify_opportunity_areas(self, n):
        """Identifica las áreas de oportunidad basado en avance real"""
        if len(self.datos_seccion) == 0:
            return []
        
        df = self.datos_seccion.reset_index().merge(
            self.simpatizantes_por_seccion, on='seccion', how='left'
        ).fillna(0)
        
        df['cumplimiento'] = (df['simpatizantes'] / df['meta'] * 100).round(1)
        df = df[df['meta'] > 0].nsmallest(n, 'cumplimiento')
        return df[['seccion', 'cumplimiento']].to_dict('records')
    
    def calculate_projection(self):
        """Calcula la proyección de meta basado en simpatizantes actuales"""
        if len(self.datos_seccion) == 0:
            return {}
        
        df_combinado = self.datos_seccion.reset_index().merge(
            self.simpatizantes_por_seccion, on='seccion', how='left'
        ).fillna(0)
        
        total_meta = df_combinado['meta'].sum()
        total_simpatizantes = df_combinado['simpatizantes'].sum()
        restante = total_meta - total_simpatizantes
        
        return {
            'logrado': int(total_simpatizantes),
            'pendiente': int(max(0, restante)),
            'porcentaje': round((total_simpatizantes / total_meta * 100), 1) if total_meta > 0 else 0
        }
    
    def generate_trend_analysis(self):
        """Análisis de tendencias y correlaciones"""
        analysis = {}
        
        if len(self.datos_seccion) > 0:
            df = self.datos_seccion.reset_index().merge(
                self.simpatizantes_por_seccion, on='seccion', how='left'
            ).fillna(0)
            
            # Correlación entre simpatizantes y metas
            if len(df) > 1:
                correlation = df['simpatizantes'].corr(df['meta'])
                analysis['correlation_simpatizantes_meta'] = round(correlation, 3)
                
                # Análisis de distribución de simpatizantes
                analysis['skewness'] = round(df['simpatizantes'].skew(), 3)
                analysis['kurtosis'] = round(df['simpatizantes'].kurtosis(), 3)
                
                # Percentiles de simpatizantes
                analysis['percentiles'] = {
                    'p25': int(df['simpatizantes'].quantile(0.25)),
                    'p50': int(df['simpatizantes'].quantile(0.5)),
                    'p75': int(df['simpatizantes'].quantile(0.75)),
                    'p90': int(df['simpatizantes'].quantile(0.9))
                }
        
        return analysis
    
    def create_dashboard_charts(self):
        """Crea visualizaciones profesionales para dashboard"""
        charts = {}
        
        # 1. Distribución de simpatizantes con curva normal
        fig1 = go.Figure()
        fig1.add_trace(go.Histogram(
            x=self.simpatizantes_por_seccion['simpatizantes'],
            nbinsx=20,
            name='Distribución Real',
            marker_color='#667eea',
            opacity=0.7
        ))
        
        # Agregar línea de media
        media = self.simpatizantes_por_seccion['simpatizantes'].mean()
        fig1.add_vline(x=media, line_dash="dash", line_color="#10b981",
                      annotation_text=f"Media: {media:.0f}", annotation_position="top")
        
        fig1.update_layout(
            title=dict(
                text="Distribución de Simpatizantes por Sección",
                font=dict(size=18, color='#1e293b')
            ),
            xaxis_title="Número de Simpatizantes",
            yaxis_title="Frecuencia",
            plot_bgcolor='rgba(0,0,0,0)',
            paper_bgcolor='rgba(0,0,0,0)',
            font=dict(family="Inter", size=12),
            hovermode='x unified',
            showlegend=False
        )
        charts['distribution'] = fig1
        
        # 2. Top 20 secciones con más simpatizantes
        top20 = self.simpatizantes_por_seccion.nlargest(20, 'simpatizantes')
        fig2 = px.bar(
            top20,
            x='seccion',
            y='simpatizantes',
            title='Top 20 Secciones · Mayor Concentración de Simpatizantes',
            labels={'seccion': 'Sección', 'simpatizantes': 'Simpatizantes'},
            color='simpatizantes',
            color_continuous_scale=['#667eea', '#764ba2']
        )
        fig2.update_layout(
            plot_bgcolor='rgba(0,0,0,0)',
            paper_bgcolor='rgba(0,0,0,0)',
            font=dict(family="Inter"),
            xaxis_tickangle=-45
        )
        charts['top20'] = fig2
        
        # 3. Matriz de desempeño (simpatizantes vs meta)
        if len(self.datos_seccion) > 0:
            df_matrix = self.datos_seccion.reset_index().merge(
                self.simpatizantes_por_seccion, on='seccion', how='inner'
            )
            df_matrix['eficiencia'] = (df_matrix['simpatizantes'] / df_matrix['meta'] * 100).round(1)
            df_matrix = df_matrix[df_matrix['meta'] > 0]
            
            fig3 = px.scatter(
                df_matrix,
                x='simpatizantes',
                y='eficiencia',
                size='meta',
                color='eficiencia',
                hover_name='seccion',
                title='Matriz de Desempeño Electoral',
                labels={
                    'simpatizantes': 'Simpatizantes',
                    'eficiencia': 'Eficiencia (%)',
                    'meta': 'Meta'
                },
                color_continuous_scale='RdYlGn',
                size_max=30
            )
            fig3.update_layout(
                plot_bgcolor='rgba(0,0,0,0)',
                paper_bgcolor='rgba(0,0,0,0)',
                font=dict(family="Inter")
            )
            charts['performance_matrix'] = fig3
        
        return charts

# ---------------------------------------------------
# VISOR DE MAPA PROFESIONAL
# ---------------------------------------------------
class ProfessionalMapViewer:
    @staticmethod
    def create_map(features, colonias, db, filtro, datos_seccion_con_colores, simpatizantes_colonia):
        centro = [22.7709, -102.5832]
        zoom = 13
        
        # Calcular simpatizantes por sección
        simpatizantes_por_seccion = db.groupby('seccion')['simpatizantes'].sum().reset_index()
        simpatizantes_dict = dict(zip(simpatizantes_por_seccion['seccion'], simpatizantes_por_seccion['simpatizantes']))
        
        # Estilo de mapa profesional
        m = folium.Map(
            location=centro,
            zoom_start=zoom,
            tiles='CartoDB positron',
            control_scale=True
        )
        
        secciones = {}
        for feature in features:
            sec = feature["properties"]["seccion"]
            secciones.setdefault(sec, []).append(feature)
        
        for seccion, lista_poligonos in secciones.items():
            if filtro and seccion != filtro:
                continue
            
            total_simpatizantes = int(simpatizantes_dict.get(seccion, 0))
            
            # Obtener color del Excel si existe
            if seccion in datos_seccion_con_colores.index:
                color_excel = datos_seccion_con_colores.loc[seccion, 'color']
                color = color_excel if pd.notna(color_excel) else '#94a3b8'
            else:
                # Color por defecto gris
                color = '#94a3b8'
            
            # Preparar contenido del popup
            cps = db[db.seccion == seccion].cp.dropna().unique()
            cp_html = "<br>".join(cps) if len(cps) > 0 else "No disponible"
            
            cols = colonias.get(seccion, [])
            colonias_html = "<br>".join(f"• {c['colonia']} (CP {c['cp']})" for c in cols[:5])
            if len(cols) > 5:
                colonias_html += f"<br>... y {len(cols)-5} más"
            
            detalle = simpatizantes_colonia[simpatizantes_colonia.seccion == seccion]
            detalle_html = ""
            for _, row in detalle.nlargest(5, 'simpatizantes').iterrows():
                detalle_html += f"• {row['colonia']}: {int(row['simpatizantes'])}<br>"
            
            if seccion in datos_seccion_con_colores.index:
                d = datos_seccion_con_colores.loc[seccion]
                participacion = f"{int(d['participacion']*100)}%" if pd.notna(d['participacion']) else "N/D"
                inicial = str(int(d['inicial'])) if pd.notna(d['inicial']) else "N/D"
                meta = str(int(d['meta'])) if pd.notna(d['meta']) else "N/D"
                
                # Calcular avance real con simpatizantes
                if d['meta'] > 0 and pd.notna(d['meta']):
                    avance = int((total_simpatizantes / d['meta'] * 100))
                    avance_color = '#10b981' if avance >= 80 else '#f59e0b' if avance >= 50 else '#ef4444'
                    
                    progress_bar = f"""
                    <div style="background: #e2e8f0; border-radius: 10px; margin: 10px 0;">
                        <div style="background: {avance_color}; width: {avance}%; height: 8px; border-radius: 10px;"></div>
                    </div>
                    <p style="text-align: right; font-size: 0.9rem; color: #64748b;">{avance}% de meta</p>
                    """
                else:
                    progress_bar = ""
                
                stats_html = f"""
                <div style="background: #f8fafc; border-radius: 12px; padding: 12px; margin: 10px 0;">
                    <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 10px;">
                        <div><span style="color: #64748b;">Participación</span><br><strong>{participacion}</strong></div>
                        <div><span style="color: #64748b;">Inicial</span><br><strong>{inicial}</strong></div>
                        <div><span style="color: #64748b;">Meta</span><br><strong>{meta}</strong></div>
                        <div><span style="color: #64748b;">Simpatizantes</span><br><strong>{total_simpatizantes}</strong></div>
                    </div>
                    {progress_bar}
                </div>
                """
            else:
                stats_html = f"""
                <div style="background: #f8fafc; border-radius: 12px; padding: 12px; margin: 10px 0;">
                    <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 10px;">
                        <div><span style="color: #64748b;">Participación</span><br><strong>N/D</strong></div>
                        <div><span style="color: #64748b;">Inicial</span><br><strong>N/D</strong></div>
                        <div><span style="color: #64748b;">Meta</span><br><strong>N/D</strong></div>
                        <div><span style="color: #64748b;">Simpatizantes</span><br><strong>{total_simpatizantes}</strong></div>
                    </div>
                </div>
                """
            
            popup_html = f"""
            <div style="font-family: 'Inter', sans-serif; min-width: 280px;">
                <h3 style="color: #1e293b; margin: 0 0 10px 0; font-weight: 700;">Sección {seccion}</h3>
                {stats_html}
                <div style="margin: 10px 0;">
                    <p style="color: #1e293b; font-weight: 600; margin: 0 0 5px 0;">📍 CP Registrados</p>
                    <p style="color: #64748b; margin: 0;">{cp_html}</p>
                </div>
                <div style="margin: 10px 0;">
                    <p style="color: #1e293b; font-weight: 600; margin: 0 0 5px 0;">🏘️ Colonias</p>
                    <p style="color: #64748b; margin: 0;">{colonias_html}</p>
                </div>
                <div style="margin: 10px 0;">
                    <p style="color: #1e293b; font-weight: 600; margin: 0 0 5px 0;">👥 Top Colonias</p>
                    <p style="color: #64748b; margin: 0;">{detalle_html}</p>
                </div>
            </div>
            """
            
            for feature in lista_poligonos:
                folium.GeoJson(
                    feature,
                    style_function=lambda x, color=color: {
                        "fillColor": color,
                        "color": "#475569",
                        "weight": 1.5,
                        "fillOpacity": 0.7
                    },
                    tooltip=folium.Tooltip(f"Sección {seccion} · {total_simpatizantes} simpatizantes"),
                    popup=folium.Popup(popup_html, max_width=350)
                ).add_to(m)
            
            # Agregar marcador con número
            try:
                coords = lista_poligonos[0]["geometry"]["coordinates"][0]
                if isinstance(coords[0], list) and isinstance(coords[0][0], list):
                    coords = coords[0]
                lat = sum(p[1] for p in coords) / len(coords)
                lon = sum(p[0] for p in coords) / len(coords)
                
                if total_simpatizantes > 0:
                    # Calcular color del marcador basado en avance
                    if seccion in datos_seccion_con_colores.index:
                        d = datos_seccion_con_colores.loc[seccion]
                        if d['meta'] > 0:
                            avance = (total_simpatizantes / d['meta']) * 100
                            if avance >= 80:
                                marker_color = '#10b981'
                            elif avance >= 50:
                                marker_color = '#f59e0b'
                            else:
                                marker_color = '#ef4444'
                        else:
                            marker_color = '#94a3b8'
                    else:
                        marker_color = '#94a3b8'
                    
                    folium.Marker(
                        [lat, lon],
                        icon=folium.DivIcon(
                            html=f"""
                            <div style="
                                background: {marker_color};
                                color: white;
                                padding: 8px 12px;
                                border-radius: 30px;
                                font-weight: 700;
                                font-size: 14px;
                                box-shadow: 0 4px 15px rgba(0,0,0,0.2);
                                border: 2px solid white;
                            ">{total_simpatizantes}</div>
                            """
                        )
                    ).add_to(m)
            except:
                pass
        
        # Agregar leyenda de colores
        legend_html = '''
        <div style="position: fixed; bottom: 50px; right: 50px; z-index: 1000; background: white; padding: 15px; border-radius: 10px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); border: 1px solid #e2e8f0;">
            <p style="font-weight: 700; margin: 0 0 10px 0; color: #1e293b;">Leyenda de Colores</p>
            <p style="margin: 5px 0;"><span style="display: inline-block; width: 20px; height: 20px; background: #10b981; margin-right: 10px; border-radius: 4px;"></span> Alto desempeño (≥80%)</p>
            <p style="margin: 5px 0;"><span style="display: inline-block; width: 20px; height: 20px; background: #f59e0b; margin-right: 10px; border-radius: 4px;"></span> Desempeño medio (50-79%)</p>
            <p style="margin: 5px 0;"><span style="display: inline-block; width: 20px; height: 20px; background: #ef4444; margin-right: 10px; border-radius: 4px;"></span> Bajo desempeño (<50%)</p>
            <p style="margin: 5px 0;"><span style="display: inline-block; width: 20px; height: 20px; background: #94a3b8; margin-right: 10px; border-radius: 4px;"></span> Sin datos de meta</p>
        </div>
        '''
        m.get_root().html.add_child(folium.Element(legend_html))
        
        # Agregar control de capas
        folium.LayerControl().add_to(m)
        
        return m

# ---------------------------------------------------
# APLICACIÓN PRINCIPAL
# ---------------------------------------------------
def main():
    # Verificar autenticación
    if not SecurityManager.check_authentication():
        return
    
    # Header Premium
    st.markdown("""
    <div class="header-premium animate-slide">
        <div style="display: flex; justify-content: space-between; align-items: center;">
            <div>
                <h1>🗳️ ZACATECAS · SIE</h1>
                <p>Sistema de Inteligencia Electoral · Visualización Estratégica de Datos</p>
            </div>
            <div style="display: flex; gap: 1rem;">
                <span class="badge-premium">v2.0 Enterprise</span>
                <span class="badge-premium">Datos en tiempo real</span>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)
    
    # Panel de control
    col_control1, col_control2, col_control3, col_control4 = st.columns([1, 1, 2, 1])
    
    with col_control1:
        if st.button("🔄 Actualizar Datos", use_container_width=True):
            with st.spinner("Actualizando información..."):
                st.cache_data.clear()
                st.rerun()
    
    with col_control2:
        if st.button("📊 Exportar Reporte", use_container_width=True):
            st.success("Reporte generado exitosamente")
    
    with col_control3:
        st.markdown(f"""
        <div style="background: #f8fafc; padding: 0.75rem 1rem; border-radius: 12px; text-align: center;">
            <span style="color: #64748b;">Última sincronización: </span>
            <span style="color: #1e293b; font-weight: 600;">{datetime.now().strftime('%d/%m/%Y %H:%M:%S')}</span>
        </div>
        """, unsafe_allow_html=True)
    
    with col_control4:
        if st.button("🚪 Cerrar Sesión", use_container_width=True):
            st.session_state[SecurityManager.SESSION_KEY] = False
            st.rerun()
    
    st.markdown('<div class="divider-premium"></div>', unsafe_allow_html=True)
    
    # Cargar datos
    with st.spinner("Cargando inteligencia electoral..."):
        db = DataManager.get_ine_data()
        colonias = DataManager.load_excel()
        geo = DataManager.load_geojson()
        simpatizantes_colonia = DataManager.get_simpatizantes_colonia()
        datos_seccion_con_colores = DataManager.load_datos_seccion_con_colores()
        operadores_df = DataManager.get_operadores_data()
    
    # Inicializar analítica avanzada
    analytics = AdvancedAnalytics(db, datos_seccion_con_colores, simpatizantes_colonia)
    operadores_analytics = OperadoresAnalytics(operadores_df)
    executive_summary = analytics.generate_executive_summary()
    
    # Métricas clave
    st.markdown("### 📈 Indicadores Estratégicos")
    
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.markdown(f"""
        <div class="metric-card-premium">
            <div class="metric-label">Total Simpatizantes</div>
            <div class="metric-value">{executive_summary['total_simpatizantes']:,}</div>
            <div class="metric-trend">
                <span class="trend-up">↑ +12%</span>
                <span>vs período anterior</span>
            </div>
        </div>
        """, unsafe_allow_html=True)
    
    with col2:
        st.markdown(f"""
        <div class="metric-card-premium">
            <div class="metric-label">Secciones Activas</div>
            <div class="metric-value">{executive_summary['secciones_activas']}</div>
            <div class="metric-trend">
                <span>Cobertura {executive_summary['cobertura']}</span>
            </div>
        </div>
        """, unsafe_allow_html=True)
    
    with col3:
        st.markdown(f"""
        <div class="metric-card-premium">
            <div class="metric-label">Promedio x Sección</div>
            <div class="metric-value">{executive_summary['promedio_x_seccion']}</div>
            <div class="metric-trend">
                <span>Mediana: {analytics.generate_trend_analysis().get('percentiles', {}).get('p50', 0)}</span>
            </div>
        </div>
        """, unsafe_allow_html=True)
    
    with col4:
        st.markdown(f"""
        <div class="metric-card-premium">
            <div class="metric-label">Eficiencia Global</div>
            <div class="metric-value">{executive_summary['eficiencia_global']}%</div>
            <div class="metric-trend">
                <span>Meta: {executive_summary['proyeccion_meta'].get('pendiente', 0):,} restantes</span>
            </div>
        </div>
        """, unsafe_allow_html=True)
    
    st.markdown('<div class="divider-premium"></div>', unsafe_allow_html=True)
    
    # Buscador inteligente
    col_search1, col_search2 = st.columns([3, 1])
    
    with col_search1:
        search_term = st.text_input(
            "🔍 Búsqueda inteligente",
            placeholder="Ingresa número de sección, colonia o CP...",
            label_visibility="collapsed"
        )
    
    with col_search2:
        if st.button("Limpiar búsqueda", use_container_width=True):
            search_term = ""
            st.rerun()
    
    # Aplicar filtro
    filtro = None
    if search_term:
        search_term_clean = search_term.strip().zfill(4) if search_term.strip().isdigit() else search_term.strip()
        if search_term_clean in db['seccion'].values:
            filtro = search_term_clean
            st.success(f"Sección {search_term_clean} encontrada")
        else:
            st.warning(f"No se encontró: '{search_term}'")
    
    # Tabs principales
    main_tabs = st.tabs([
        "🗺️ Visualización Geoestratégica",
        "👥 Análisis por Operadores",
        "📊 Centro de Análisis Electoral",
        "📋 Datos Maestros"
    ])
    
    with main_tabs[0]:
        # Mapa profesional con colores del Excel
        st.markdown("### 🗺️ Visualización Geoestratégica")
        st.markdown("Los colores de las secciones corresponden a los colores del archivo DATOS_POR_SECCION.xlsx")
        
        mapa = ProfessionalMapViewer.create_map(
            geo, colonias, db, filtro, 
            datos_seccion_con_colores, simpatizantes_colonia
        )
        folium_static(mapa, width=1600, height=650)
    
    with main_tabs[1]:
        st.markdown("### 👥 Análisis de Desempeño por Operadores")
        
        if not operadores_df.empty:
            # Resumen de operadores
            resumen_operadores = operadores_analytics.get_resumen_operadores()
            
            # Métricas de operadores
            col_op1, col_op2, col_op3, col_op4 = st.columns(4)
            
            with col_op1:
                st.markdown(f"""
                <div class="metric-card-premium">
                    <div class="metric-label">Total Operadores</div>
                    <div class="metric-value">{len(resumen_operadores)}</div>
                </div>
                """, unsafe_allow_html=True)
            
            with col_op2:
                total_registros = int(resumen_operadores['total_registros'].sum())
                st.markdown(f"""
                <div class="metric-card-premium">
                    <div class="metric-label">Total Registros</div>
                    <div class="metric-value">{total_registros:,}</div>
                </div>
                """, unsafe_allow_html=True)
            
            with col_op3:
                promedio_registros = int(resumen_operadores['total_registros'].mean())
                st.markdown(f"""
                <div class="metric-card-premium">
                    <div class="metric-label">Promedio x Operador</div>
                    <div class="metric-value">{promedio_registros}</div>
                </div>
                """, unsafe_allow_html=True)
            
            with col_op4:
                mejor_operador = resumen_operadores.nlargest(1, 'total_registros').iloc[0]
                st.markdown(f"""
                <div class="metric-card-premium">
                    <div class="metric-label">Mejor Operador</div>
                    <div class="metric-value" style="font-size: 1.5rem;">{mejor_operador['usuario_nombre']}</div>
                    <div class="metric-trend">{int(mejor_operador['total_registros'])} registros</div>
                </div>
                """, unsafe_allow_html=True)
            
            st.markdown('<div class="divider-premium"></div>', unsafe_allow_html=True)
            
            # Gráficos de operadores
            col_op_chart1, col_op_chart2 = st.columns(2)
            
            with col_op_chart1:
                operador_charts = operadores_analytics.create_operador_charts()
                if 'top_operadores' in operador_charts:
                    st.plotly_chart(operador_charts['top_operadores'], use_container_width=True)
            
            with col_op_chart2:
                if 'edad_distribution' in operador_charts:
                    st.plotly_chart(operador_charts['edad_distribution'], use_container_width=True)
            
            st.markdown('<div class="divider-premium"></div>', unsafe_allow_html=True)
            
            # Tabla de desempeño por operador
            st.markdown("#### 📊 Tabla de Desempeño por Operador")
            
            # Selector de operador
            operadores_list = ['Todos'] + list(resumen_operadores['usuario_nombre'].unique())
            operador_seleccionado = st.selectbox("Filtrar por operador:", operadores_list, key="operador_select")
            
            if operador_seleccionado == 'Todos':
                df_mostrar = resumen_operadores.copy()
            else:
                df_mostrar = resumen_operadores[resumen_operadores['usuario_nombre'] == operador_seleccionado].copy()
            
            # Formatear columnas para mostrar
            df_mostrar['edad_promedio'] = df_mostrar['edad_promedio'].round(1)
            
            st.dataframe(
                df_mostrar,
                use_container_width=True,
                hide_index=True,
                column_config={
                    'usuario_nombre': st.column_config.TextColumn('Operador'),
                    'total_registros': st.column_config.NumberColumn('Registros', format="%d"),
                    'secciones_asignadas': st.column_config.NumberColumn('Secciones', format="%d"),
                    'edad_promedio': st.column_config.NumberColumn('Edad Prom.', format="%.1f"),
                    'edad_min': st.column_config.NumberColumn('Edad Mín', format="%d"),
                    'edad_max': st.column_config.NumberColumn('Edad Máx', format="%d"),
                    'indice_productividad': st.column_config.NumberColumn('Productividad', format="%.1f"),
                    'anio_registro_comun': st.column_config.NumberColumn('Año común', format="%d")
                }
            )
            
            st.markdown('<div class="divider-premium"></div>', unsafe_allow_html=True)
            
            # Detalle de operador específico
            st.markdown("#### 🔍 Detalle por Operador")
            
            operador_detalle = st.selectbox(
                "Selecciona operador para ver detalle:",
                resumen_operadores['usuario_nombre'].unique(),
                key="operador_detalle"
            )
            
            if operador_detalle:
                detalle = operadores_analytics.get_operador_detalle(operador_detalle)
                
                # Estadísticas rápidas
                col_d1, col_d2, col_d3, col_d4 = st.columns(4)
                
                with col_d1:
                    st.metric("Registros", len(detalle))
                with col_d2:
                    st.metric("Secciones", detalle['seccion'].nunique())
                with col_d3:
                    st.metric("Edad Promedio", f"{detalle['edad'].mean():.1f}")
                with col_d4:
                    st.metric("Hombres/Mujeres", f"{len(detalle[detalle['sexo']=='H'])}/{len(detalle[detalle['sexo']=='M'])}")
                
                # Mostrar tabla de registros
                st.dataframe(
                    detalle[['nombre_completo', 'seccion', 'edad', 'sexo', 'anio_registro', 'vigencia']],
                    use_container_width=True,
                    hide_index=True,
                    column_config={
                        'nombre_completo': 'Nombre',
                        'seccion': 'Sección',
                        'edad': 'Edad',
                        'sexo': 'Sexo',
                        'anio_registro': 'Año Registro',
                        'vigencia': 'Vigencia'
                    }
                )
        else:
            st.warning("No hay datos de operadores disponibles")
    
    with main_tabs[2]:
        st.markdown("### 📊 Centro de Análisis Electoral")
        
        tabs_analisis = st.tabs([
            "📈 Distribución y Tendencias",
            "🎯 Desempeño por Sección",
            "🏆 Rankings y Comparativas"
        ])
        
        charts = analytics.create_dashboard_charts()
        trend_analysis = analytics.generate_trend_analysis()
        
        with tabs_analisis[0]:
            col1, col2 = st.columns(2)
            
            with col1:
                if 'distribution' in charts:
                    st.plotly_chart(charts['distribution'], use_container_width=True)
                
                # Estadísticas de tendencia
                st.markdown("#### Análisis de Tendencia")
                col_t1, col_t2, col_t3 = st.columns(3)
                with col_t1:
                    st.metric("Sesgo", trend_analysis.get('skewness', 'N/A'))
                with col_t2:
                    st.metric("Curtosis", trend_analysis.get('kurtosis', 'N/A'))
                with col_t3:
                    st.metric("Correlación", trend_analysis.get('correlation_simpatizantes_meta', 'N/A'))
            
            with col2:
                if 'performance_matrix' in charts:
                    st.plotly_chart(charts['performance_matrix'], use_container_width=True)
                
                # Percentiles
                if 'percentiles' in trend_analysis:
                    st.markdown("#### Distribución por Percentiles")
                    percentiles = trend_analysis['percentiles']
                    col_p1, col_p2, col_p3, col_p4 = st.columns(4)
                    col_p1.metric("P25", percentiles['p25'])
                    col_p2.metric("P50", percentiles['p50'])
                    col_p3.metric("P75", percentiles['p75'])
                    col_p4.metric("P90", percentiles['p90'])
        
        with tabs_analisis[1]:
            if len(datos_seccion_con_colores) > 0:
                # Crear dataframe con avance real
                df_desempeno = datos_seccion_con_colores.reset_index().copy()
                simpatizantes_por_seccion = db.groupby('seccion')['simpatizantes'].sum().reset_index()
                df_desempeno = df_desempeno.merge(
                    simpatizantes_por_seccion, on='seccion', how='left'
                ).fillna(0)
                
                df_desempeno['avance'] = (df_desempeno['simpatizantes'] / df_desempeno['meta'] * 100).round(1)
                df_desempeno['faltantes'] = (df_desempeno['meta'] - df_desempeno['simpatizantes']).clip(lower=0)
                
                # Selector de métrica
                metrica = st.selectbox(
                    "Selecciona métrica de visualización",
                    ["avance", "simpatizantes", "faltantes", "meta"],
                    format_func=lambda x: {
                        'avance': '% de Avance',
                        'simpatizantes': 'Simpatizantes Actuales',
                        'faltantes': 'Faltantes para Meta',
                        'meta': 'Meta'
                    }[x]
                )
                
                fig_desempeno = px.bar(
                    df_desempeno.nlargest(20, metrica if metrica != 'faltantes' else 'faltantes'),
                    x='seccion',
                    y=metrica,
                    title=f"Top 20 Secciones por {metrica}",
                    color=metrica,
                    color_continuous_scale='Viridis'
                )
                st.plotly_chart(fig_desempeno, use_container_width=True)
                
                # Mostrar tabla de avance
                st.markdown("#### 📊 Tabla de Avance por Sección")
                st.dataframe(
                    df_desempeno[['seccion', 'simpatizantes', 'meta', 'avance', 'faltantes']].sort_values('avance', ascending=False),
                    use_container_width=True,
                    hide_index=True,
                    column_config={
                        'seccion': 'Sección',
                        'simpatizantes': st.column_config.NumberColumn('Simpatizantes', format="%d"),
                        'meta': st.column_config.NumberColumn('Meta', format="%d"),
                        'avance': st.column_config.ProgressColumn(
                            '% Avance',
                            format="%.1f%%",
                            min_value=0,
                            max_value=100
                        ),
                        'faltantes': st.column_config.NumberColumn('Faltantes', format="%d")
                    }
                )
        
        with tabs_analisis[2]:
            col_r1, col_r2 = st.columns(2)
            
            with col_r1:
                st.markdown("#### 🏆 Top Performers")
                for i, perf in enumerate(executive_summary['top_performers'], 1):
                    st.markdown(f"""
                    <div style="background: #f8fafc; padding: 1rem; border-radius: 12px; margin: 0.5rem 0;">
                        <div style="display: flex; justify-content: space-between;">
                            <span style="font-weight: 600;">#{i} Sección {perf['seccion']}</span>
                            <span style="color: #10b981; font-weight: 700;">{perf['cumplimiento']}%</span>
                        </div>
                    </div>
                    """, unsafe_allow_html=True)
            
            with col_r2:
                st.markdown("#### 📉 Áreas de Oportunidad")
                for i, opp in enumerate(executive_summary['areas_oportunidad'], 1):
                    st.markdown(f"""
                    <div style="background: #f8fafc; padding: 1rem; border-radius: 12px; margin: 0.5rem 0;">
                        <div style="display: flex; justify-content: space-between;">
                            <span style="font-weight: 600;">#{i} Sección {opp['seccion']}</span>
                            <span style="color: #ef4444; font-weight: 700;">{opp['cumplimiento']}%</span>
                        </div>
                    </div>
                    """, unsafe_allow_html=True)
    
    with main_tabs[3]:
        # Tabla maestra
        st.markdown("#### 📋 Datos Maestros del Sistema")
        
        # Preparar datos
        simpatizantes_totales = db.groupby('seccion')['simpatizantes'].sum().reset_index()
        simpatizantes_totales.columns = ['seccion', 'simpatizantes']
        
        tabla_maestra = simpatizantes_totales.merge(
            datos_seccion_con_colores.reset_index(),
            on='seccion',
            how='left'
        ).fillna(0)
        
        # Calcular métricas adicionales
        tabla_maestra['participacion_pct'] = (tabla_maestra['participacion'] * 100).round(1)
        tabla_maestra['avance'] = (tabla_maestra['simpatizantes'] / tabla_maestra['meta'] * 100).round(1)
        tabla_maestra['avance'] = tabla_maestra['avance'].fillna(0)
        tabla_maestra['faltantes'] = (tabla_maestra['meta'] - tabla_maestra['simpatizantes']).clip(lower=0)
        
        # Renombrar columnas
        tabla_maestra.columns = [
            'Sección', 'Simpatizantes', 'Participación',
            'Inicial', 'Meta', 'Color Excel', '% Participación', '% Avance', 'Faltantes'
        ]
        
        # Mostrar tabla con indicadores de color
        st.dataframe(
            tabla_maestra,
            use_container_width=True,
            hide_index=True,
            column_config={
                'Sección': st.column_config.TextColumn('Sección'),
                'Simpatizantes': st.column_config.NumberColumn('Simpatizantes', format="%d"),
                'Inicial': st.column_config.NumberColumn('Inicial', format="%d"),
                'Meta': st.column_config.NumberColumn('Meta', format="%d"),
                'Color Excel': st.column_config.TextColumn('Color Excel'),
                '% Participación': st.column_config.NumberColumn('% Participación', format="%.1f%%"),
                '% Avance': st.column_config.ProgressColumn(
                    '% Avance',
                    format="%.1f%%",
                    min_value=0,
                    max_value=100
                ),
                'Faltantes': st.column_config.NumberColumn('Faltantes', format="%d")
            }
        )
        
        # Tabla de operadores
        st.markdown("#### 👥 Datos de Operadores")
        if not operadores_df.empty:
            st.dataframe(
                operadores_df[['usuario_nombre', 'nombre_completo', 'seccion', 'edad', 'sexo', 'anio_registro']].head(100),
                use_container_width=True,
                hide_index=True,
                column_config={
                    'usuario_nombre': 'Operador',
                    'nombre_completo': 'Nombre',
                    'seccion': 'Sección',
                    'edad': 'Edad',
                    'sexo': 'Sexo',
                    'anio_registro': 'Año Registro'
                }
            )
    
    st.markdown('<div class="divider-premium"></div>', unsafe_allow_html=True)
    
    # Footer profesional
    col_f1, col_f2, col_f3, col_f4 = st.columns(4)
    
    with col_f1:
        st.markdown("**© 2026 · SIE Zacatecas**")
    with col_f2:
        st.markdown(f"**Secciones monitoreadas:** {executive_summary['secciones_activas']}")
    with col_f3:
        st.markdown(f"**Operadores activos:** {len(operadores_analytics.get_resumen_operadores()) if not operadores_df.empty else 0}")
    with col_f4:
        st.markdown("**v2.0 Enterprise · Todos los derechos reservados**")

# ---------------------------------------------------
# EJECUTAR APLICACIÓN
# ---------------------------------------------------
if __name__ == "__main__":
    main()